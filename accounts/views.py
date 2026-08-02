from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth import (
    authenticate,
    login,
    logout,
    update_session_auth_hash,
)
from django.db.models import Sum, F
from django.db.models.functions import TruncMonth
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import (
    make_password,
    check_password,
)

from accounts.decorators import subscription_required
from accounts.models import Subscription

from packages.models import (
    Trip,
    Customer,
    Expense,
    AgencySettings,
)

from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from urllib.parse import quote

from django.http import HttpResponse

from openpyxl import Workbook

import os
import arabic_reshaper
from bidi.algorithm import get_display

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.enums import TA_RIGHT, TA_CENTER
from reportlab.lib.styles import (
    getSampleStyleSheet,
    ParagraphStyle,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
def login_view(request):

    if request.method == "POST":

        email = request.POST["email"]
        password = request.POST["password"]

        user = authenticate(
            request,
            username=email,
            password=password
        )

        if user is not None:
            login(request, user)
            return redirect("/dashboard/")

        return render(request, "login.html", {
            "error": "البريد الإلكتروني أو كلمة المرور غير صحيحة."
        })

    return render(request, "login.html")


def signup_view(request):

    if request.method == "POST":

        agency_name = request.POST["agency_name"]
        email = request.POST["email"]
        password = request.POST["password"]

        if User.objects.filter(username=email).exists():
            return render(request, "signup.html", {
                "error": "البريد الإلكتروني مستخدم بالفعل."
            })

        user = User.objects.create_user(
            username=email,
            email=email,
            password=password,
            first_name=agency_name,
        )

        Subscription.objects.create(
            user=user
        )

        login(request, user)

        return redirect("/dashboard/")

    return render(request, "signup.html")
@login_required
@subscription_required
def dashboard_view(request):
    from decimal import Decimal
    from django.db.models import Sum
    from django.utils import timezone

    request.session.pop("financial_access", None)

    subscription = Subscription.objects.get(user=request.user)

    # حساب الأيام المتبقية
    remaining_days = subscription.days_left

    # إذا كان الاشتراك مدى الحياة
    if remaining_days != "مدى الحياة" and remaining_days < 0:
        remaining_days = 0

    trips = Trip.objects.filter(user=request.user).order_by("-id")
    customers = Customer.objects.filter(user=request.user)

    trips_count = trips.count()
    customers_count = customers.count()

    total_received = customers.aggregate(
        total=Sum("amount_paid")
    )["total"] or Decimal("0")

    total_remaining = sum(
        (customer.remaining_amount() for customer in customers),
        Decimal("0"),
    )

    latest_customers = customers.order_by("-id")[:5]
    latest_trips = trips.order_by("-id")[:5]

    trip_seats = []
    total_available_seats = 0
    low_seats_trips = []
    full_trips = []

    for trip in trips:
        booked_customers = customers.filter(trip=trip).count()

        available_seats = max(
            trip.seats - booked_customers,
            0,
        )

        trip_seats.append({
            "trip": trip,
            "available_seats": available_seats,
            "booked_customers": booked_customers,
        })

        total_available_seats += available_seats

        if available_seats == 0:
            full_trips.append(trip)

        elif available_seats <= 5:
            low_seats_trips.append(trip)

    unpaid_customers = customers.exclude(
        amount_paid__gte=F("total_price")
    )

    double_rooms = customers.filter(room_type="ثنائية").count()
    triple_rooms = customers.filter(room_type="ثلاثية").count()
    quad_rooms = customers.filter(room_type="رباعية").count()
    quint_rooms = customers.filter(room_type="خماسية").count()

    unread_support = SupportMessage.objects.filter(
        user=request.user,
        is_admin=True,
        is_read=False,
    ).count()

    today = timezone.now().date()

    today_bookings = customers.filter(
        created_at=today
    ).count()

    month_bookings = customers.filter(
        created_at__year=today.year,
        created_at__month=today.month,
    ).count()

    full_trips_count = len(full_trips)

    paid_customers = customers.filter(
        amount_paid__gte=F("total_price")
    ).count()

    remaining_customers = customers.filter(
        amount_paid__lt=F("total_price")
    ).count()

    total_capacity = trips.aggregate(
        total=Sum("seats")
    )["total"] or 0

    return render(
        request,
        "dashboard.html",
        {
            "subscription": subscription,
            "remaining_days": remaining_days,
            "trips_count": trips_count,
            "customers_count": customers_count,
            "total_received": total_received,
            "total_remaining": total_remaining,
            "latest_customers": latest_customers,
            "latest_trips": latest_trips,
            "low_seats_trips": low_seats_trips,
            "full_trips": full_trips,
            "full_trips_count": full_trips_count,
            "unpaid_customers": unpaid_customers,
            "double_rooms": double_rooms,
            "triple_rooms": triple_rooms,
            "quad_rooms": quad_rooms,
            "quint_rooms": quint_rooms,
            "unread_support": unread_support,
            "today_bookings": today_bookings,
            "month_bookings": month_bookings,
            "total_seats": total_available_seats,
            "total_capacity": total_capacity,
            "paid_customers": paid_customers,
            "remaining_customers": remaining_customers,
            "trip_seats": trip_seats,
        },
    )
def financial_period(request):
    today = date.today()

    preset = request.GET.get("preset", "")

    start_date_str = request.GET.get("start_date", "")
    end_date_str = request.GET.get("end_date", "")

    start_date = None
    end_date = None

    if preset == "this_month":
        start_date = today.replace(day=1)
        end_date = today

    elif preset == "last_month":
        first_this_month = today.replace(day=1)
        end_date = first_this_month - timedelta(days=1)
        start_date = end_date.replace(day=1)

    elif preset == "this_year":
        start_date = date(today.year, 1, 1)
        end_date = today

    else:
        if start_date_str:
            try:
                start_date = date.fromisoformat(start_date_str)
            except ValueError:
                start_date = None

        if end_date_str:
            try:
                end_date = date.fromisoformat(end_date_str)
            except ValueError:
                end_date = None

    return start_date, end_date
def pdf_ar(text):
    text = str(text)
    reshaped = arabic_reshaper.reshape(text)
    return get_display(reshaped)
@login_required
@subscription_required
def export_financial_excel(request):

    # ---------------------------------
    # حماية التقرير المالي
    # ---------------------------------
    if not request.session.get("financial_access"):
        return redirect("financial_unlock")

    # ---------------------------------
    # اسم الوكالة
    # ---------------------------------
    agency_settings = AgencySettings.objects.filter(
        user=request.user
    ).first()

    agency_name = (
        agency_settings.agency_name.strip()
        if agency_settings
        and agency_settings.agency_name
        else "الوكالة"
    )

    # ---------------------------------
    # الفترة المالية
    # ---------------------------------
    start_date, end_date = financial_period(request)

    # ---------------------------------
    # بيانات الوكالة الحالية فقط
    # ---------------------------------
    customers = Customer.objects.filter(
        user=request.user
    )

    expenses = Expense.objects.filter(
        user=request.user
    ).select_related("trip")

    # ---------------------------------
    # تطبيق الفترة المالية
    # ---------------------------------
    if start_date:

        customers = customers.filter(
            created_at__gte=start_date
        )

        expenses = expenses.filter(
            date__gte=start_date
        )

    if end_date:

        customers = customers.filter(
            created_at__lte=end_date
        )

        expenses = expenses.filter(
            date__lte=end_date
        )

    # ---------------------------------
    # الرحلات المرتبطة فعليًا
    # بالحجوزات والمصاريف داخل الفترة
    # ---------------------------------
    customer_trip_ids = set(
        customers
        .exclude(trip_id__isnull=True)
        .values_list(
            "trip_id",
            flat=True
        )
    )

    expense_trip_ids = set(
        expenses
        .exclude(trip_id__isnull=True)
        .values_list(
            "trip_id",
            flat=True
        )
    )

    relevant_trip_ids = (
        customer_trip_ids
        | expense_trip_ids
    )

    trips = Trip.objects.filter(
        user=request.user,
        id__in=relevant_trip_ids
    ).order_by("-id")
    # ---------------------------------
    # الحسابات المالية
    # ---------------------------------
    total_sales = (
        customers.aggregate(
            total=Sum("total_price")
        )["total"]
        or Decimal("0")
    )

    total_received = (
        customers.aggregate(
            total=Sum("amount_paid")
        )["total"]
        or Decimal("0")
    )

    total_remaining = (
        total_sales - total_received
    )

    total_expenses = (
        expenses.aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0")
    )

    net_cash = (
        total_received - total_expenses
    )

    estimated_profit = (
        total_sales - total_expenses
    )

    # ---------------------------------
    # إنشاء ملف Excel
    # ---------------------------------
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "الملخص المالي"

    # ---------------------------------
    # عنوان التقرير
    # ---------------------------------
    sheet["A1"] = (
        f"التقرير المالي - {agency_name}"
    )

    # ---------------------------------
    # الفترة
    # ---------------------------------
    if start_date or end_date:

        period_text = (
            "الفترة: "
            f"{start_date or 'بداية غير محددة'} "
            "إلى "
            f"{end_date or 'نهاية غير محددة'}"
        )

    else:

        period_text = "الفترة: كل السجلات"

    sheet["A2"] = period_text

    # ---------------------------------
    # المؤشرات المالية
    # ---------------------------------
    sheet["A4"] = "المؤشر"
    sheet["B4"] = "القيمة (دج)"

    sheet["A5"] = "إجمالي المبيعات"
    sheet["B5"] = float(total_sales)

    sheet["A6"] = "المبلغ المحصل"
    sheet["B6"] = float(total_received)

    sheet["A7"] = "المبلغ المتبقي"
    sheet["B7"] = float(total_remaining)

    sheet["A8"] = "إجمالي المصاريف"
    sheet["B8"] = float(total_expenses)

    sheet["A9"] = "السيولة الصافية"
    sheet["B9"] = float(net_cash)

    sheet["A10"] = "الربح التقديري"
    sheet["B10"] = float(estimated_profit)

    # ---------------------------------
    # تنسيق الملخص
    # ---------------------------------
    for cell in sheet["A"]:
        cell.font = cell.font.copy(
            bold=True
        )

    for cell in sheet[4]:
        cell.font = cell.font.copy(
            bold=True
        )

    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 22
    # ---------------------------------
    # ورقة الرحلات
    # ---------------------------------
    trip_sheet = workbook.create_sheet(
        "الرحلات"
    )

    headers = [
        "الرحلة",
        "إجمالي المبيعات",
        "المحصل",
        "المتبقي",
        "المصاريف",
        "الربح التقديري",
    ]

    for col, header in enumerate(
        headers,
        start=1
    ):

        cell = trip_sheet.cell(
            row=1,
            column=col
        )

        cell.value = header

        cell.font = cell.font.copy(
            bold=True
        )

    row = 2

    for trip in trips:

        trip_customers = customers.filter(
            trip=trip
        )

        trip_expenses = expenses.filter(
            trip=trip
        )

        trip_sales = (
            trip_customers.aggregate(
                total=Sum("total_price")
            )["total"]
            or Decimal("0")
        )

        trip_received = (
            trip_customers.aggregate(
                total=Sum("amount_paid")
            )["total"]
            or Decimal("0")
        )

        trip_expenses_total = (
            trip_expenses.aggregate(
                total=Sum("amount")
            )["total"]
            or Decimal("0")
        )

        trip_remaining = (
            trip_sales - trip_received
        )

        trip_profit = (
            trip_sales - trip_expenses_total
        )

        trip_sheet.cell(
            row=row,
            column=1
        ).value = trip.name

        trip_sheet.cell(
            row=row,
            column=2
        ).value = float(trip_sales)

        trip_sheet.cell(
            row=row,
            column=3
        ).value = float(trip_received)

        trip_sheet.cell(
            row=row,
            column=4
        ).value = float(trip_remaining)

        trip_sheet.cell(
            row=row,
            column=5
        ).value = float(
            trip_expenses_total
        )

        trip_sheet.cell(
            row=row,
            column=6
        ).value = float(trip_profit)

        row += 1

    # عرض الأعمدة
    trip_widths = {
        "A": 28,
        "B": 20,
        "C": 20,
        "D": 20,
        "E": 20,
        "F": 20,
    }

    for column, width in trip_widths.items():
        trip_sheet.column_dimensions[
            column
        ].width = width

    # ---------------------------------
    # ورقة المصاريف
    # ---------------------------------
    expense_sheet = workbook.create_sheet(
        "المصاريف"
    )

    expense_headers = [
        "المصروف",
        "الفئة",
        "الرحلة",
        "المبلغ",
        "التاريخ",
        "الملاحظات",
    ]

    for col, header in enumerate(
        expense_headers,
        start=1
    ):

        cell = expense_sheet.cell(
            row=1,
            column=col
        )

        cell.value = header

        cell.font = cell.font.copy(
            bold=True
        )

    row = 2

    for expense in expenses:

        expense_sheet.cell(
            row=row,
            column=1
        ).value = expense.title

        expense_sheet.cell(
            row=row,
            column=2
        ).value = expense.category

        expense_sheet.cell(
            row=row,
            column=3
        ).value = (
            expense.trip.name
            if expense.trip
            else "مصروف عام"
        )

        expense_sheet.cell(
            row=row,
            column=4
        ).value = float(
            expense.amount
        )

        expense_sheet.cell(
            row=row,
            column=5
        ).value = expense.date.strftime(
            "%Y-%m-%d"
        )

        expense_sheet.cell(
            row=row,
            column=6
        ).value = expense.notes or ""

        row += 1

    expense_widths = {
        "A": 28,
        "B": 20,
        "C": 28,
        "D": 20,
        "E": 15,
        "F": 40,
    }

    for column, width in expense_widths.items():
        expense_sheet.column_dimensions[
            column
        ].width = width
        # ---------------------------------
    # إعداد اسم الملف
    # ---------------------------------
    filename = (
        f"التقرير المالي_{agency_name}.xlsx"
    )

    # ---------------------------------
    # إرسال الملف
    # ---------------------------------
    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
    )

    response["Content-Disposition"] = (
        "attachment; "
        f"filename*=UTF-8''{quote(filename)}"
    )

    return response
@login_required
@subscription_required
def export_financial_pdf(request):

    # ---------------------------------
    # حماية التقرير المالي
    # ---------------------------------
    if not request.session.get("financial_access"):
        return redirect("financial_unlock")

    # ---------------------------------
    # اسم الوكالة
    # ---------------------------------
    agency_settings = AgencySettings.objects.filter(
        user=request.user
    ).first()

    agency_name = (
        agency_settings.agency_name.strip()
        if agency_settings and agency_settings.agency_name
        else "الوكالة"
    )

    # ---------------------------------
    # الفترة المالية
    # ---------------------------------
    start_date, end_date = financial_period(request)

    # ---------------------------------
    # بيانات الوكالة الحالية فقط
    # ---------------------------------
    customers = Customer.objects.filter(
        user=request.user
    )

    expenses = Expense.objects.filter(
        user=request.user
    ).select_related("trip")

    trips = Trip.objects.filter(
        user=request.user
    ).order_by("-id")

    # ---------------------------------
    # تطبيق الفترة
    # ---------------------------------
    if start_date:
        customers = customers.filter(
            created_at__gte=start_date
        )

        expenses = expenses.filter(
            date__gte=start_date
        )

    if end_date:
        customers = customers.filter(
            created_at__lte=end_date
        )

        expenses = expenses.filter(
            date__lte=end_date
        )

    # ---------------------------------
    # الحسابات المالية
    # ---------------------------------
    total_sales = (
        customers.aggregate(
            total=Sum("total_price")
        )["total"]
        or Decimal("0")
    )

    total_received = (
        customers.aggregate(
            total=Sum("amount_paid")
        )["total"]
        or Decimal("0")
    )

    total_remaining = (
        total_sales - total_received
    )

    total_expenses = (
        expenses.aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0")
    )

    net_cash = (
        total_received - total_expenses
    )

    estimated_profit = (
        total_sales - total_expenses
    )

    # ---------------------------------
    # العثور على خط يدعم العربية
    # ---------------------------------
    font_candidates = [
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\Arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    ]

    font_path = next(
        (
            path
            for path in font_candidates
            if os.path.exists(path)
        ),
        None,
    )

    if not font_path:
        return HttpResponse(
            "لم يتم العثور على خط يدعم العربية.",
            status=500,
        )

    pdfmetrics.registerFont(
        TTFont(
            "TravoraArabic",
            font_path,
        )
    )

    # ---------------------------------
    # إنشاء PDF
    # ---------------------------------
    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=35,
        leftMargin=35,
        topMargin=35,
        bottomMargin=35,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "TravoraTitle",
        parent=styles["Title"],
        fontName="TravoraArabic",
        fontSize=20,
        leading=28,
        alignment=TA_CENTER,
        spaceAfter=12,
    )

    section_style = ParagraphStyle(
        "TravoraSection",
        parent=styles["Heading2"],
        fontName="TravoraArabic",
        fontSize=13,
        leading=20,
        alignment=TA_RIGHT,
        spaceAfter=10,
    )

    normal_style = ParagraphStyle(
        "TravoraNormal",
        parent=styles["Normal"],
        fontName="TravoraArabic",
        fontSize=10,
        leading=17,
        alignment=TA_RIGHT,
    )
    small_style = ParagraphStyle(
        "TravoraSmall",
        parent=styles["Normal"],
        fontName="TravoraArabic",
        fontSize=8,
        leading=13,
        alignment=TA_RIGHT,
    )

    story = []

    # ---------------------------------
    # عنوان التقرير
    # ---------------------------------
    story.append(
        Paragraph(
            pdf_ar(
    f"التقرير المالي — {agency_name}"
),
            title_style,
        )
    )

    # ---------------------------------
    # تاريخ إنشاء التقرير
    # ---------------------------------
    report_date = date.today()

    story.append(
        Paragraph(
            pdf_ar(
                "تاريخ إنشاء التقرير: "
                + report_date.strftime("%d-%m-%Y")
            ),
            normal_style,
        )
    )

    # ---------------------------------
    # الفترة
    # ---------------------------------
    if start_date or end_date:
        period_text = (
            "الفترة: "
            + str(start_date or "بداية غير محددة")
            + " إلى "
            + str(end_date or "نهاية غير محددة")
        )
    else:
        period_text = "الفترة: كل السجلات"

    story.append(
        Paragraph(
            pdf_ar(period_text),
            normal_style,
        )
    )

    story.append(
        Spacer(1, 18)
    )

    # ---------------------------------
    # ملخص الوضع المالي
    # ---------------------------------
    story.append(
        Paragraph(
            pdf_ar("ملخص الوضع المالي"),
            section_style,
        )
    )

    summary_data = [
        [
            pdf_ar("المؤشر"),
            pdf_ar("القيمة"),
        ],
        [
            pdf_ar("إجمالي المبيعات"),
            f"{total_sales} دج",
        ],
        [
            pdf_ar("المبلغ المحصل"),
            f"{total_received} دج",
        ],
        [
            pdf_ar("المبلغ المتبقي"),
            f"{total_remaining} دج",
        ],
        [
            pdf_ar("إجمالي المصاريف"),
            f"{total_expenses} دج",
        ],
        [
            pdf_ar("السيولة الصافية"),
            f"{net_cash} دج",
        ],
        [
            pdf_ar("الربح التقديري"),
            f"{estimated_profit} دج",
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[260, 180],
    )

    summary_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#1769aa"),
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white,
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, -1),
                "TravoraArabic",
            ),
            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "RIGHT",
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#d1d5db"),
            ),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#f7fafc"),
                ],
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                9,
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                9,
            ),
        ])
    )

    story.append(summary_table)

    story.append(
        Spacer(1, 25)
    )

    # ---------------------------------
    # الربح حسب الرحلة
    # ---------------------------------
    story.append(
        Paragraph(
            pdf_ar("تفاصيل الربح حسب الرحلة"),
            section_style,
        )
    )
    trip_data = [
        [
            pdf_ar("الرحلة"),
            pdf_ar("المبيعات"),
            pdf_ar("المحصل"),
            pdf_ar("المتبقي"),
            pdf_ar("المصاريف"),
            pdf_ar("الربح"),
        ]
    ]

    for trip in trips:

        trip_customers = customers.filter(
            trip=trip
        )

        trip_expenses = expenses.filter(
            trip=trip
        )

        sales = (
            trip_customers.aggregate(
                total=Sum("total_price")
            )["total"]
            or Decimal("0")
        )

        received = (
            trip_customers.aggregate(
                total=Sum("amount_paid")
            )["total"]
            or Decimal("0")
        )

        trip_exp = (
            trip_expenses.aggregate(
                total=Sum("amount")
            )["total"]
            or Decimal("0")
        )

        remaining = (
            sales - received
        )

        profit = (
            sales - trip_exp
        )

        trip_data.append([
            pdf_ar(trip.name),
            f"{sales} دج",
            f"{received} دج",
            f"{remaining} دج",
            f"{trip_exp} دج",
            f"{profit} دج",
        ])

    if len(trip_data) == 1:
        trip_data.append([
            pdf_ar("لا توجد رحلات ضمن الفترة"),
            "0 دج",
            "0 دج",
            "0 دج",
            "0 دج",
            "0 دج",
        ])

    trip_table = Table(
        trip_data,
        repeatRows=1,
        colWidths=[
            120,
            80,
            80,
            80,
            80,
            80,
        ],
    )

    trip_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#1769aa"),
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white,
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, -1),
                "TravoraArabic",
            ),
            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "RIGHT",
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#d1d5db"),
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8,
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                7,
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                7,
            ),
        ])
    )

    story.append(trip_table)

    story.append(
        Spacer(1, 25)
    )

    # ---------------------------------
    # المصاريف المسجلة
    # ---------------------------------
    story.append(
        Paragraph(
            pdf_ar("تفاصيل المصاريف المسجلة"),
            section_style,
        )
    )

    expense_data = [
        [
            pdf_ar("المصروف"),
            pdf_ar("الفئة"),
            pdf_ar("الرحلة"),
            pdf_ar("المبلغ"),
            pdf_ar("التاريخ"),
        ]
    ]

    for expense in expenses:

        expense_data.append([
            pdf_ar(expense.title),
            pdf_ar(expense.category),
            pdf_ar(
                expense.trip.name
                if expense.trip
                else "مصروف عام"
            ),
            f"{expense.amount} دج",
            expense.date.strftime("%Y-%m-%d"),
        ])

    if len(expense_data) == 1:
        expense_data.append([
            pdf_ar("لا توجد مصاريف مسجلة"),
            "-",
            "-",
            "0 دج",
            "-",
        ])

    expense_table = Table(
        expense_data,
        repeatRows=1,
        colWidths=[
            120,
            85,
            120,
            80,
            85,
        ],
    )
    expense_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#ef4444"),
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white,
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, -1),
                "TravoraArabic",
            ),
            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "RIGHT",
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor("#d1d5db"),
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8,
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                7,
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                7,
            ),
        ])
    )

    story.append(expense_table)

    # ---------------------------------
    # خاتمة التقرير
    # ---------------------------------
    story.append(
        Spacer(1, 30)
    )

    story.append(
        Paragraph(
            pdf_ar(
                f"هذا التقرير خاص بـ {agency_name} "
                "وتم إنشاؤه بواسطة Travora."
            ),
            small_style,
        )
    )

    story.append(
        Spacer(1, 8)
    )

    story.append(
        Paragraph(
            pdf_ar("Travora ©"),
            normal_style,
        )
    )

    # ---------------------------------
    # إنشاء الملف وإرساله
    # ---------------------------------
    document.build(story)

    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/pdf",
    )

    response["Content-Disposition"] = (
        'attachment; filename="travora_financial_report.pdf"'
    )

    return response
@login_required
@subscription_required
def profit_loss(request):

    if not request.session.get("financial_access"):
        return redirect("financial_unlock")

    # ---------------------------------
    # التاريخ الحالي
    # ---------------------------------
    today = date.today()

    # ---------------------------------
    # الفترات السريعة
    # ---------------------------------
    preset = request.GET.get("preset", "")

    start_date_str = request.GET.get(
        "start_date",
        ""
    )

    end_date_str = request.GET.get(
        "end_date",
        ""
    )

    start_date = None
    end_date = None

    if preset == "this_month":

        start_date = today.replace(day=1)
        end_date = today

        start_date_str = start_date.isoformat()
        end_date_str = end_date.isoformat()

    elif preset == "last_month":

        first_this_month = today.replace(day=1)

        end_date = (
            first_this_month
            - timedelta(days=1)
        )

        start_date = end_date.replace(
            day=1
        )

        start_date_str = start_date.isoformat()
        end_date_str = end_date.isoformat()

    elif preset == "this_year":

        start_date = date(
            today.year,
            1,
            1
        )

        end_date = today

        start_date_str = start_date.isoformat()
        end_date_str = end_date.isoformat()

    else:

        if start_date_str:

            try:
                start_date = date.fromisoformat(
                    start_date_str
                )
            except ValueError:
                start_date = None

        if end_date_str:

            try:
                end_date = date.fromisoformat(
                    end_date_str
                )
            except ValueError:
                end_date = None

    # ---------------------------------
    # بيانات الوكالة الحالية فقط
    # ---------------------------------
    customers = Customer.objects.filter(
        user=request.user
    )

    expenses = Expense.objects.filter(
        user=request.user
    ).select_related("trip")

    trips = Trip.objects.filter(
        user=request.user
    ).order_by("-id")

    # ---------------------------------
    # تطبيق الفترة
    # ---------------------------------
    if start_date:

        customers = customers.filter(
            created_at__gte=start_date
        )

        expenses = expenses.filter(
            date__gte=start_date
        )

    if end_date:

        customers = customers.filter(
            created_at__lte=end_date
        )

        expenses = expenses.filter(
            date__lte=end_date
        )

    # ---------------------------------
    # إضافة مصروف
    # ---------------------------------
    if request.method == "POST":

        trip_id = request.POST.get("trip")

        trip = None

        if trip_id:

            trip = Trip.objects.get(
                id=trip_id,
                user=request.user
            )

        Expense.objects.create(
            user=request.user,
            trip=trip,
            title=request.POST["title"],
            category=request.POST["category"],
            amount=request.POST["amount"],
            notes=request.POST.get(
                "notes",
                ""
            ),
        )

        return redirect("profit_loss")

    # ---------------------------------
    # الحسابات الرئيسية
    # ---------------------------------
    total_sales = (
        customers.aggregate(
            total=Sum("total_price")
        )["total"]
        or Decimal("0")
    )

    total_received = (
        customers.aggregate(
            total=Sum("amount_paid")
        )["total"]
        or Decimal("0")
    )

    total_remaining = (
        total_sales
        - total_received
    )

    total_expenses = (
        expenses.aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0")
    )

    net_cash = (
        total_received
        - total_expenses
    )
    estimated_profit = (
        total_sales
        - total_expenses
    )

    # ---------------------------------
    # ربح كل رحلة
    # ---------------------------------
    trip_reports = []

    for trip in trips:

        trip_customers = customers.filter(
            trip=trip
        )

        trip_expenses = expenses.filter(
            trip=trip
        )

        trip_sales = (
            trip_customers.aggregate(
                total=Sum("total_price")
            )["total"]
            or Decimal("0")
        )

        trip_received = (
            trip_customers.aggregate(
                total=Sum("amount_paid")
            )["total"]
            or Decimal("0")
        )

        trip_expenses_total = (
            trip_expenses.aggregate(
                total=Sum("amount")
            )["total"]
            or Decimal("0")
        )

        trip_remaining = (
            trip_sales
            - trip_received
        )

        trip_profit = (
            trip_sales
            - trip_expenses_total
        )

        trip_reports.append({
            "trip": trip,
            "sales": trip_sales,
            "received": trip_received,
            "remaining": trip_remaining,
            "expenses": trip_expenses_total,
            "profit": trip_profit,
        })

    # ---------------------------------
    # التقرير المالي الشهري
    # ---------------------------------
    monthly_sales = (
        customers
        .annotate(
            month=TruncMonth(
                "created_at"
            )
        )
        .values("month")
        .annotate(
            total=Sum("total_price")
        )
        .order_by("month")
    )

    monthly_expenses = (
        expenses
        .annotate(
            month=TruncMonth(
                "date"
            )
        )
        .values("month")
        .annotate(
            total=Sum("amount")
        )
        .order_by("month")
    )

    monthly_data = {}

    arabic_months = {
        1: "جانفي",
        2: "فيفري",
        3: "مارس",
        4: "أفريل",
        5: "ماي",
        6: "جوان",
        7: "جويلية",
        8: "أوت",
        9: "سبتمبر",
        10: "أكتوبر",
        11: "نوفمبر",
        12: "ديسمبر",
    }

    # المبيعات الشهرية
    for item in monthly_sales:

        if item["month"]:

            key = item["month"].strftime(
                "%Y-%m"
            )

            monthly_data[key] = {
                "month": key,
                "sales": (
                    item["total"]
                    or Decimal("0")
                ),
                "expenses": Decimal("0"),
            }
# ---------------------------------
    # المصاريف الشهرية
    # ---------------------------------
    for item in monthly_expenses:

        if item["month"]:

            key = item["month"].strftime(
                "%Y-%m"
            )

            if key not in monthly_data:

                monthly_data[key] = {
                    "month": key,
                    "sales": Decimal("0"),
                    "expenses": Decimal("0"),
                }

            monthly_data[key]["expenses"] = (
                item["total"]
                or Decimal("0")
            )

    # ---------------------------------
    # التقرير الشهري
    # ---------------------------------
    monthly_report = []

    for item in monthly_data.values():

        item["profit"] = (
            item["sales"]
            - item["expenses"]
        )

        year, month_number = map(
            int,
            item["month"].split("-")
        )

        item["year"] = year

        item["month_number"] = month_number

        item["month_ar"] = (
            arabic_months.get(
                month_number,
                item["month"]
            )
        )

        monthly_report.append(item)

    monthly_report.sort(
        key=lambda item: item["month"]
    )

    # ---------------------------------
    # الملخص المالي الذكي
    # ---------------------------------

    # أفضل شهر نحسبوه فقط من الأشهر
    # اللي فيها مبيعات فعلية
    months_with_sales = [
        item
        for item in monthly_report
        if item["sales"] > Decimal("0")
    ]

    best_month = None

    if months_with_sales:

        best_month = max(
            months_with_sales,
            key=lambda item: item["profit"]
        )

    # ---------------------------------
    # أفضل رحلة
    # ---------------------------------
    best_trip = None

    if trip_reports:

        best_trip = max(
            trip_reports,
            key=lambda item: item["profit"]
        )

    # ---------------------------------
    # عرض الصفحة
    # ---------------------------------
    return render(
        request,
        "profit_loss.html",
        {
            "expenses": expenses,

            "trips": trips,

            "trip_reports": trip_reports,

            "monthly_report": monthly_report,

            "total_sales": total_sales,

            "total_received": total_received,

            "total_remaining": total_remaining,

            "total_expenses": total_expenses,

            "net_cash": net_cash,

            "estimated_profit": estimated_profit,

            "best_month": best_month,

            "best_trip": best_trip,

            "start_date": start_date_str,

            "end_date": end_date_str,

            "today": today,

            "preset": preset,
        },
    )
def logout_view(request):
    logout(request)
    return redirect("/")
def subscription_expired(request):

    return render(
        request,
        "subscription_expired.html"
    )
def subscription_plans(request):

    return render(
        request,
        "subscription_plans.html"
    )
def payment_info(request):
    return render(request, "payment_info.html")
@login_required
def change_password(request):

    if request.method == "POST":

        current_password = request.POST["current_password"]
        new_password = request.POST["new_password"]
        confirm_password = request.POST["confirm_password"]

        if not request.user.check_password(current_password):
            return render(request, "change_password.html", {
                "error": "كلمة المرور الحالية غير صحيحة."
            })

        if new_password != confirm_password:
            return render(request, "change_password.html", {
                "error": "كلمتا المرور غير متطابقتين."
            })

        request.user.set_password(new_password)
        request.user.save()

        update_session_auth_hash(request, request.user)

        return render(request, "change_password.html", {
            "success": "تم تغيير كلمة المرور بنجاح."
        })

    return render(request, "change_password.html")
def payment_page(request):
    return render(request, "payment.html")
from .models import SupportMessage
from django.contrib.auth.decorators import login_required
@login_required
def support_chat(request):

    if request.method == "POST":
        SupportMessage.objects.create(
            user=request.user,
            message=request.POST.get("message"),
            image=request.FILES.get("image"),
            is_admin=False,
        )

        return redirect("support_chat")

    messages = SupportMessage.objects.filter(
        user=request.user
    ).order_by("created_at")

    # تعليم رسائل الإدارة كمقروءة
    messages.filter(
        is_admin=True,
        is_read=False
    ).update(is_read=True)

    return render(
        request,
        "support_chat.html",
        {
            "messages": messages,
        },
    )
from django.contrib.auth.decorators import login_required
from django.shortcuts import render

@login_required
def payment_page(request):
    return render(request, "payment.html")
@login_required
@subscription_required
def set_financial_password(request):
    subscription = Subscription.objects.get(user=request.user)

    if request.method == "POST":
        password = request.POST.get("password", "")
        confirm_password = request.POST.get("confirm_password", "")

        if not password:
            return render(
                request,
                "set_financial_password.html",
                {"error": "يرجى إدخال كلمة السر."},
            )

        if len(password) < 6:
            return render(
                request,
                "set_financial_password.html",
                {"error": "كلمة السر يجب أن تكون 6 أحرف على الأقل."},
            )

        if password != confirm_password:
            return render(
                request,
                "set_financial_password.html",
                {"error": "كلمتا السر غير متطابقتين."},
            )

        subscription.financial_password = make_password(password)
        subscription.save(update_fields=["financial_password"])

        return redirect("financial_unlock")

    return render(request, "set_financial_password.html")
@login_required
@subscription_required
def financial_unlock(request):
    subscription = Subscription.objects.get(user=request.user)

    if request.method == "POST":
        password = request.POST.get("password", "")

        if (
            subscription.financial_password
            and check_password(
                password,
                subscription.financial_password
            )
        ):
            request.session["financial_access"] = True
            return redirect("profit_loss")

        return render(
            request,
            "financial_unlock.html",
            {"error": "كلمة السر المالية غير صحيحة."},
        )

    return render(request, "financial_unlock.html")
@login_required
@subscription_required
def subscription_chat(request):
    return render(request, "subscription_chat.html")
@login_required
def admin_panel(request):
    if not request.user.is_superuser:
        return redirect("dashboard")

    if not request.session.get("admin_panel_access"):
        return redirect("admin_panel_login")

    return render(request, "admin_panel.html")
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from .models import PaymentInfoRequest, Subscription

@login_required
def subscription_requests(request):

    if request.method == "POST":
        request_id = request.POST.get("request_id")

        payment_request = get_object_or_404(
            PaymentInfoRequest,
            id=request_id
        )

        payment_request.approved = True
        payment_request.save()

        subscription, created = Subscription.objects.get_or_create(
            user=payment_request.user
        )

        subscription.lifetime = True
        subscription.save()

        return redirect("subscription_requests")

    requests = PaymentInfoRequest.objects.all().order_by("-created_at")

    return render(
        request,
        "subscription_requests.html",
        {
            "requests": requests
        }
    )
from .models import SupportMessage

@login_required
def support_messages(request):

    messages = SupportMessage.objects.select_related("user").order_by("-created_at")

    return render(
        request,
        "support_messages.html",
        {
            "messages": messages,
        },
    )
@login_required
def agencies(request):
    agencies = User.objects.all().order_by("-id")
    return render(
        request,
        "agencies.html",
        {
            "agencies": agencies,
        },
    )
@login_required
def admin_statistics(request):
    return render(request, "admin_statistics.html")
from django.contrib.auth.hashers import check_password
from django.shortcuts import render, redirect

ADMIN_PASSWORD = "اكتبي_هنا_كلمة_سر_قوية"

@login_required
def admin_panel_login(request):
    if not request.user.is_superuser:
        return redirect("dashboard")

    if request.method == "POST":
        password = request.POST.get("password")

        if password == ADMIN_PASSWORD:
            request.session["admin_panel_access"] = True
            return redirect("admin_panel")

        return render(request, "admin_panel_login.html", {
            "error": "كلمة المرور غير صحيحة."
        })

    return render(request, "admin_panel_login.html")